#!/usr/bin/env python3
"""从 Excel 导入真实菜品数据，替换假数据"""
import sqlite3
import openpyxl
import re
import os

DB_PATH = os.path.join(os.path.dirname(__file__), 'database.db')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'data', '北交食堂菜品收集终版.xlsx')


def parse_price(val):
    """价格清洗：'6 元' → 6.0, '大份 20 元 / 中份 15 元 / 小份 10 元' → 10.0"""
    if val is None:
        return 0.0
    s = str(val).strip()
    nums = re.findall(r'(\d+\.?\d*)', s)
    if not nums:
        return 0.0
    return float(nums[0])


def parse_tags(val):
    """标签清洗，返回逗号分隔的字符串"""
    if val is None:
        return ''
    return str(val).strip().replace('、', ',').replace('，', ',')


def main():
    print('读取 Excel...')
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]

    # 解析 Excel 行
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name:
            continue
        name = str(name).strip()
        # 新 Excel 列: 菜品名称 | 单位 | 价格 | 所属食堂 | 楼层 | 所属档口 | 标签 | 供餐时段
        unit = str(row[1]).strip() if row[1] else ''
        price_raw = str(row[2]).strip() if row[2] else '0'
        canteen_name = str(row[3]).strip() if row[3] else ''
        floor = str(row[4]).strip() if row[4] else ''
        stall_name = str(row[5]).strip() if row[5] else ''
        tags = parse_tags(row[6])
        meal_time = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        is_new = False  # 新表无此列，统一为非新品
        category = ''   # 稍后由 fill_categories.py 填充

        rows.append({
            'name': name,
            'price': parse_price(price_raw),
            'canteen_name': canteen_name,
            'floor': floor,
            'stall_name': stall_name,
            'tags': tags,
            'is_new': is_new,
            'category': category,
        })

    print(f'共读取 {len(rows)} 道菜品')

    # 统计食堂+楼层+档口组合
    canteen_floor_set = set()
    stall_set = set()
    for r in rows:
        canteen_floor_set.add((r['canteen_name'], r['floor']))
        stall_set.add((r['canteen_name'], r['floor'], r['stall_name']))

    print(f'食堂楼层: {len(canteen_floor_set)} 个')
    print(f'档口: {len(stall_set)} 个')

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 清除旧数据（保留用户表）
    print('清除旧菜品数据...')
    cursor.execute('DELETE FROM checkin')
    cursor.execute('DELETE FROM comment')
    cursor.execute('DELETE FROM rating')
    cursor.execute('DELETE FROM post')
    cursor.execute('DELETE FROM dish')
    cursor.execute('DELETE FROM stall')
    cursor.execute('DELETE FROM canteen')

    # 插入食堂
    canteen_map = {}
    for idx, (name, floor) in enumerate(sorted(canteen_floor_set), 1):
        desc = ''
        if '东区' in name or '明湖' in name:
            desc = '东校区主食堂'
        elif '活动中心' in name:
            desc = '学生活动中心餐厅'
        cursor.execute(
            'INSERT INTO canteen (id, name, floor, crowd_status, description) VALUES (?,?,?,?,?)',
            (idx, name, floor, '一般', desc)
        )
        canteen_map[(name, floor)] = idx

    # 插入档口
    stall_map = {}
    stall_id_counter = 1
    for canteen_name, floor, stall_name in sorted(stall_set):
        c_id = canteen_map[(canteen_name, floor)]
        cursor.execute(
            'INSERT INTO stall (id, canteen_id, name) VALUES (?,?,?)',
            (stall_id_counter, c_id, stall_name)
        )
        stall_map[(canteen_name, floor, stall_name)] = stall_id_counter
        stall_id_counter += 1

    # 插入菜品
    print('插入菜品...')
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for idx, r in enumerate(rows, 1):
        s_id = stall_map[(r['canteen_name'], r['floor'], r['stall_name'])]
        create_time = now if r['is_new'] else '2026-01-01 00:00:00'
        cursor.execute(
            'INSERT INTO dish (id, stall_id, name, price, tags, is_new, category, create_time) VALUES (?,?,?,?,?,?,?,?)',
            (idx, s_id, r['name'], r['price'], r['tags'], 1 if r['is_new'] else 0, r['category'], create_time)
        )

    conn.commit()
    conn.close()

    print(f'\n导入完成！')
    print(f'  食堂楼层: {len(canteen_floor_set)} 个')
    print(f'  档口: {len(stall_set)} 个')
    print(f'  菜品: {len(rows)} 道')
    print(f'  用户数据已保留')


if __name__ == '__main__':
    main()
